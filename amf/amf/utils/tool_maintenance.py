# -*- coding: utf-8 -*-
# Copyright (c) 2026, libracore AG and contributors
# For license information, please see license.txt

from __future__ import unicode_literals

import os
import re
import unicodedata
from datetime import date, datetime
from difflib import SequenceMatcher

import frappe
from frappe import _
from frappe.custom.doctype.custom_field.custom_field import create_custom_fields
from frappe.utils import add_days, add_months, add_years, cint, cstr, getdate, nowdate


TOOL_ITEM_GROUP = "Tool"
ACTIVE_PLAN_STATUS = "Active"

ITEM_MAINTENANCE_FIELDS = (
	"tool_serial_number",
	"tool_equipment_type",
	"tool_ownership",
	"tool_location",
	"tool_responsible",
	"tool_required_ppe",
	"tool_calibration_procedure",
	"tool_maintenance_instructions",
)

TOOL_MAINTENANCE_CUSTOM_FIELDS = {
	"Item": [
		{
			"fieldname": "tool_maintenance_section",
			"fieldtype": "Section Break",
			"label": "Tool Maintenance",
			"insert_after": "item_defaults",
			"depends_on": "eval:doc.item_group=='Tool'",
		},
		{
			"fieldname": "tool_serial_number",
			"fieldtype": "Data",
			"label": "Equipment Serial Number",
			"insert_after": "tool_maintenance_section",
			"in_standard_filter": 1,
		},
		{
			"fieldname": "tool_equipment_type",
			"fieldtype": "Data",
			"label": "Equipment Type",
			"insert_after": "tool_serial_number",
		},
		{
			"fieldname": "tool_ownership",
			"fieldtype": "Data",
			"label": "Property / Ownership",
			"insert_after": "tool_equipment_type",
		},
		{
			"fieldname": "tool_location",
			"fieldtype": "Data",
			"label": "Location",
			"insert_after": "tool_ownership",
			"in_standard_filter": 1,
		},
		{
			"fieldname": "tool_maintenance_column_break",
			"fieldtype": "Column Break",
			"insert_after": "tool_location",
		},
		{
			"fieldname": "tool_responsible",
			"fieldtype": "Link",
			"label": "Maintenance Responsible",
			"options": "Employee",
			"insert_after": "tool_maintenance_column_break",
			"in_standard_filter": 1,
		},
		{
			"fieldname": "tool_required_ppe",
			"fieldtype": "Small Text",
			"label": "Required PPE",
			"insert_after": "tool_responsible",
		},
		{
			"fieldname": "tool_calibration_procedure",
			"fieldtype": "Small Text",
			"label": "Calibration / Verification Procedure",
			"insert_after": "tool_required_ppe",
		},
		{
			"fieldname": "tool_maintenance_instructions",
			"fieldtype": "Small Text",
			"label": "Maintenance Instructions",
			"insert_after": "tool_calibration_procedure",
		},
		{
			"fieldname": "tool_maintenance_summary_section",
			"fieldtype": "Section Break",
			"label": "Maintenance Planning Summary",
			"insert_after": "tool_maintenance_instructions",
			"depends_on": "eval:doc.item_group=='Tool'",
			"collapsible": 1,
		},
		{
			"fieldname": "tool_last_maintenance_date",
			"fieldtype": "Date",
			"label": "Last Intervention",
			"insert_after": "tool_maintenance_summary_section",
			"read_only": 1,
			"no_copy": 1,
		},
		{
			"fieldname": "tool_next_maintenance_date",
			"fieldtype": "Date",
			"label": "Next Intervention",
			"insert_after": "tool_last_maintenance_date",
			"read_only": 1,
			"no_copy": 1,
			"in_standard_filter": 1,
		},
		{
			"fieldname": "tool_maintenance_status",
			"fieldtype": "Select",
			"label": "Maintenance Status",
			"options": "\nNo Plan\nPlanned\nDue Soon\nOverdue",
			"insert_after": "tool_next_maintenance_date",
			"read_only": 1,
			"no_copy": 1,
			"in_standard_filter": 1,
		},
		{
			"fieldname": "tool_maintenance_summary_column_break",
			"fieldtype": "Column Break",
			"insert_after": "tool_maintenance_status",
		},
		{
			"fieldname": "tool_open_maintenance_plans",
			"fieldtype": "Int",
			"label": "Active Plans",
			"insert_after": "tool_maintenance_summary_column_break",
			"read_only": 1,
			"no_copy": 1,
		},
		{
			"fieldname": "tool_overdue_maintenance_plans",
			"fieldtype": "Int",
			"label": "Overdue Plans",
			"insert_after": "tool_open_maintenance_plans",
			"read_only": 1,
			"no_copy": 1,
		},
	],
}


def sync_tool_maintenance_custom_fields():
	"""Install Tool-only Item fields and refresh their planning summaries."""
	create_custom_fields(TOOL_MAINTENANCE_CUSTOM_FIELDS, update=True)
	frappe.clear_cache(doctype="Item")
	return sync_all_tool_maintenance_summaries()


def validate_tool_item(item_code):
	item = frappe.db.get_value(
		"Item",
		item_code,
		["name", "item_name", "item_group", "disabled"],
		as_dict=True,
	)
	if not item:
		frappe.throw(_("Item {0} does not exist.").format(item_code))
	if item.item_group != TOOL_ITEM_GROUP:
		frappe.throw(
			_("Maintenance plans and logs are only available for Items in the Tool Item Group.")
		)
	return item


def validate_tool_item_group_change(doc, method=None):
	"""Keep existing plans and logs attached to a Tool Item."""
	if doc.is_new() or doc.item_group == TOOL_ITEM_GROUP:
		return
	previous_group = frappe.db.get_value("Item", doc.name, "item_group")
	if previous_group != TOOL_ITEM_GROUP:
		return

	linked_plan = frappe.db.get_value("Tool Maintenance Plan", {"item_code": doc.name}, "name")
	linked_log = frappe.db.get_value("Tool Maintenance Log", {"item_code": doc.name}, "name")
	if linked_plan or linked_log:
		links = [value for value in (linked_plan, linked_log) if value]
		frappe.throw(
			_(
				"Item {0} has Tool maintenance records ({1}). Remove or reassign those records "
				"before changing its Item Group."
			).format(doc.name, ", ".join(links))
		)


def sync_tool_item_after_save(doc, method=None):
	"""Initialize and refresh maintenance summaries after every Item save."""
	values = sync_item_maintenance_summary(doc.name) or {}
	for fieldname, value in values.items():
		if doc.meta.get_field(fieldname):
			doc.set(fieldname, value)
	return values


def calculate_next_due_date(completed_on, frequency_value, frequency_unit):
	frequency_value = cint(frequency_value)
	if not completed_on or frequency_value <= 0:
		return None
	completed_on = getdate(completed_on)
	if frequency_unit == "Days":
		return add_days(completed_on, frequency_value)
	if frequency_unit == "Weeks":
		return add_days(completed_on, frequency_value * 7)
	if frequency_unit == "Months":
		return add_months(completed_on, frequency_value)
	if frequency_unit == "Years":
		return add_years(completed_on, frequency_value)
	return None


def get_plan_state(plans, today=None):
	"""Return the dynamic state shared by Item summaries and the dashboard."""
	today = getdate(today or nowdate())
	plans = list(plans or [])
	due_plans = [plan for plan in plans if plan.get("next_due_date")]
	overdue = [plan for plan in due_plans if getdate(plan.next_due_date) < today]
	due_soon = [
		plan
		for plan in due_plans
		if today <= getdate(plan.next_due_date) <= add_days(today, cint(plan.get("warning_days") or 30))
	]

	if overdue:
		status = "Overdue"
	elif due_soon:
		status = "Due Soon"
	elif plans:
		status = "Planned"
	else:
		status = "No Plan"

	next_plan = min(due_plans, key=lambda plan: getdate(plan.next_due_date)) if due_plans else None
	return {
		"status": status,
		"active_count": len(plans),
		"overdue_count": len(overdue),
		"due_soon_count": len(due_soon),
		"next_due_date": next_plan.next_due_date if next_plan else None,
		"next_plan": next_plan,
	}


def sync_item_maintenance_summary(item_code, exclude_plan=None, exclude_log=None):
	if not item_code or not frappe.db.exists("Item", item_code):
		return None
	if not _maintenance_columns_exist():
		return None

	item_group = frappe.db.get_value("Item", item_code, "item_group")
	if item_group != TOOL_ITEM_GROUP:
		values = {
			"tool_last_maintenance_date": None,
			"tool_next_maintenance_date": None,
			"tool_maintenance_status": None,
			"tool_open_maintenance_plans": 0,
			"tool_overdue_maintenance_plans": 0,
		}
	else:
		plans = frappe.get_all(
			"Tool Maintenance Plan",
			filters={"item_code": item_code, "status": ACTIVE_PLAN_STATUS},
			fields=["name", "activity", "next_due_date", "warning_days", "responsible"],
		)
		if exclude_plan:
			plans = [plan for plan in plans if plan.name != exclude_plan]
		state = get_plan_state(plans)

		log_filters = {"item_code": item_code}
		logs = frappe.get_all(
			"Tool Maintenance Log",
			filters=log_filters,
			fields=["name", "intervention_date"],
			order_by="intervention_date desc, modified desc",
			limit_page_length=20,
		)
		if exclude_log:
			logs = [log for log in logs if log.name != exclude_log]

		values = {
			"tool_last_maintenance_date": logs[0].intervention_date if logs else None,
			"tool_next_maintenance_date": state["next_due_date"],
			"tool_maintenance_status": state["status"],
			"tool_open_maintenance_plans": state["active_count"],
			"tool_overdue_maintenance_plans": state["overdue_count"],
		}

	frappe.db.set_value("Item", item_code, values, update_modified=False)
	return values


def sync_all_tool_maintenance_summaries():
	if not _maintenance_columns_exist():
		return {"updated": 0, "skipped": "missing_columns"}
	if not frappe.db.exists("DocType", "Tool Maintenance Plan"):
		return {"updated": 0, "skipped": "missing_doctypes"}

	items = frappe.get_all("Item", filters={"item_group": TOOL_ITEM_GROUP}, fields=["name"])
	for item in items:
		sync_item_maintenance_summary(item.name)
	return {"updated": len(items)}


def rebuild_plan_completion(plan_name, exclude_log=None):
	if not plan_name or not frappe.db.exists("Tool Maintenance Plan", plan_name):
		return
	plan = frappe.db.get_value(
		"Tool Maintenance Plan",
		plan_name,
		[
			"item_code", "status", "closed_on_completion", "next_due_date",
			"frequency_value", "frequency_unit",
		],
		as_dict=True,
	)
	logs = frappe.get_all(
		"Tool Maintenance Log",
		filters={"maintenance_plan": plan_name},
		fields=["name", "intervention_date", "next_due_date"],
		order_by="intervention_date desc, modified desc",
		limit_page_length=20,
	)
	if exclude_log:
		logs = [log for log in logs if log.name != exclude_log]
	latest = logs[0] if logs else None
	next_due_date = None
	if latest:
		next_due_date = latest.next_due_date or calculate_next_due_date(
			latest.intervention_date,
			plan.frequency_value,
			plan.frequency_unit,
		)
	updates = {
		"last_completed_on": latest.intervention_date if latest else None,
		"last_log": latest.name if latest else None,
	}
	if next_due_date:
		updates["next_due_date"] = next_due_date
		if cint(plan.closed_on_completion):
			updates["status"] = ACTIVE_PLAN_STATUS
			updates["closed_on_completion"] = 0
	elif latest and plan.status == ACTIVE_PLAN_STATUS:
		# A completed one-time plan keeps its original due date for the audit trail,
		# but is removed from the active planning queue.
		updates["status"] = "Closed"
		updates["closed_on_completion"] = 1
	elif not latest and cint(plan.closed_on_completion):
		# Removing or moving the completion log restores the original one-time plan.
		updates["status"] = ACTIVE_PLAN_STATUS
		updates["closed_on_completion"] = 0
	frappe.db.set_value(
		"Tool Maintenance Plan",
		plan_name,
		updates,
		update_modified=False,
	)
	sync_item_maintenance_summary(plan.item_code)
	return updates


def preview_maintenance_workbook(file_path):
	return import_maintenance_workbook(file_path=file_path, dry_run=True, commit=False)


def import_maintenance_workbook(file_path, dry_run=True, commit=False):
	"""Conservatively import the SCM equipment workbook.

	Codes whose current ERP Item name no longer resembles the spreadsheet equipment
	name are reported as conflicts and never updated. This is intentional because the
	legacy workbook contains several codes that have since been reused.
	"""
	from openpyxl import load_workbook

	dry_run = cint(dry_run)
	path = _resolve_workbook_path(file_path)
	workbook = load_workbook(path, data_only=True, read_only=True)
	items = frappe.get_all(
		"Item",
		fields=["name", "item_name", "item_group", "disabled"] + list(ITEM_MAINTENANCE_FIELDS),
		limit_page_length=5000,
	)
	items_by_code = {item.name: item for item in items}
	report = {
		"file": path,
		"dry_run": bool(dry_run),
		"item_updates": [],
		"plans": [],
		"logs": [],
		"conflicts": [],
		"skipped": [],
	}

	if "Sheet1" in workbook.sheetnames:
		_import_equipment_rows(
			workbook["Sheet1"], items, items_by_code, report, dry_run
		)

	for sheet_name in ("Maintenance & Intervention", "Maintenance Assemblage"):
		if sheet_name in workbook.sheetnames:
			_import_history_rows(
				workbook[sheet_name], items, items_by_code, report, dry_run
			)

	report["summary"] = {
		"items_updated": len(report["item_updates"]),
		"plans_created": len([row for row in report["plans"] if row.get("action") == "create"]),
		"logs_created": len([row for row in report["logs"] if row.get("action") == "create"]),
		"conflicts": len(report["conflicts"]),
		"skipped": len(report["skipped"]),
	}
	if not dry_run:
		sync_all_tool_maintenance_summaries()
		if cint(commit):
			frappe.db.commit()
	return report


def _import_equipment_rows(sheet, items, items_by_code, report, dry_run):
	for row_number, row in _worksheet_rows(sheet):
		item, reason = _resolve_item(
			row.get("Item code"), row.get("Name"), items, items_by_code
		)
		if not item:
			_bucket_resolution_issue(report, sheet.title, row_number, row, reason)
			continue

		responsible = _employee_for_name(row.get("Responsible"))
		updates = {
			"tool_serial_number": _clean_value(row.get("Serial No.")),
			"tool_equipment_type": _clean_value(row.get("Type")),
			"tool_ownership": _clean_value(row.get("Proprety")),
			"tool_location": _clean_value(row.get("Location")),
			"tool_responsible": responsible,
			"tool_required_ppe": _clean_value(row.get("EPI nécessaires")),
			"tool_calibration_procedure": _clean_value(row.get("Procédé de calibration")),
			"tool_maintenance_instructions": _clean_value(row.get("Maintenance")),
		}
		updates = {key: value for key, value in updates.items() if value not in (None, "")}
		updates = {key: value for key, value in updates.items() if item.get(key) != value}
		if not updates:
			continue
		report["item_updates"].append(
			{"item_code": item.name, "source_row": row_number, "fields": sorted(updates)}
		)
		if not dry_run:
			frappe.db.set_value("Item", item.name, updates)


def _import_history_rows(sheet, items, items_by_code, report, dry_run):
	for row_number, row in _worksheet_rows(sheet):
		item, reason = _resolve_item(
			row.get("N°article"), row.get("Equipement"), items, items_by_code
		)
		activity = _clean_value(row.get("Interventions"))
		if not item:
			_bucket_resolution_issue(report, sheet.title, row_number, row, reason)
			continue
		if not activity:
			report["skipped"].append(
				{"sheet": sheet.title, "row": row_number, "reason": "missing_intervention"}
			)
			continue

		intervention_date = parse_workbook_date(row.get("Date"))
		next_due_date = parse_workbook_date(row.get("Prochaine intervention"))
		maintenance_type = _normalize_maintenance_type(row.get("Types d'interventions"))
		responsible_name = _clean_value(row.get("Qui"))
		responsible = _employee_for_name(responsible_name)
		source_base = "SCM.3000:{0}:{1}".format(sheet.title, row_number)

		plan_name = None
		if next_due_date:
			plan_source = source_base + ":plan"
			plan_name = frappe.db.get_value(
				"Tool Maintenance Plan", {"source_reference": plan_source}, "name"
			)
			action = "existing" if plan_name else "create"
			report["plans"].append(
				{
					"action": action,
					"item_code": item.name,
					"activity": activity,
					"next_due_date": cstr(next_due_date),
					"source_row": row_number,
				}
			)
			if not dry_run and not plan_name:
				plan = frappe.get_doc(
					{
						"doctype": "Tool Maintenance Plan",
						"item_code": item.name,
						"maintenance_type": maintenance_type,
						"activity": activity,
						"status": ACTIVE_PLAN_STATUS,
						"next_due_date": next_due_date,
						"responsible": responsible,
						"notes": _clean_value(row.get("Remarques")),
						"source_reference": plan_source,
					}
				).insert(ignore_permissions=True)
				plan_name = plan.name

		if intervention_date:
			log_source = source_base + ":log"
			log_name = frappe.db.get_value(
				"Tool Maintenance Log", {"source_reference": log_source}, "name"
			)
			action = "existing" if log_name else "create"
			report["logs"].append(
				{
					"action": action,
					"item_code": item.name,
					"intervention": activity,
					"intervention_date": cstr(intervention_date),
					"source_row": row_number,
				}
			)
			if not dry_run and not log_name:
				frappe.get_doc(
					{
						"doctype": "Tool Maintenance Log",
						"item_code": item.name,
						"maintenance_plan": plan_name,
						"intervention_type": maintenance_type,
						"intervention": activity,
						"intervention_date": intervention_date,
						"responsible": responsible,
						"performed_by": responsible_name,
						"record_reference": _clean_value(row.get("Enregistrement")),
						"remarks": _clean_value(row.get("Remarques")),
						"next_due_date": next_due_date,
						"source_reference": log_source,
					}
				).insert(ignore_permissions=True)
		elif not next_due_date:
			report["skipped"].append(
				{"sheet": sheet.title, "row": row_number, "reason": "unusable_date"}
			)


def _worksheet_rows(sheet):
	rows = sheet.iter_rows(values_only=True)
	headers = [_clean_header(value) for value in next(rows)]
	for row_number, values in enumerate(rows, start=2):
		row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
		if any(value not in (None, "") for value in values):
			yield row_number, row


def _resolve_item(raw_code, equipment_name, items, items_by_code):
	codes = _item_code_candidates(raw_code)
	for code in codes:
		item = items_by_code.get(code)
		if not item:
			continue
		if item.item_group != TOOL_ITEM_GROUP:
			return None, "not_tool"
		if not _names_compatible(equipment_name, item.item_name):
			return None, "code_name_conflict"
		return item, None

	if codes:
		return None, "missing_item"

	name = _clean_value(equipment_name)
	if not name:
		return None, "missing_item_reference"
	strong_matches = []
	for item in items:
		if item.item_group != TOOL_ITEM_GROUP:
			continue
		ratio = SequenceMatcher(None, _normalize_name(name), _normalize_name(item.item_name)).ratio()
		if ratio >= 0.78:
			strong_matches.append(item)
	if len(strong_matches) == 1:
		return strong_matches[0], None
	return None, "ambiguous_name" if strong_matches else "missing_item"


def _names_compatible(source_name, current_name):
	source = _normalize_name(source_name)
	current = _normalize_name(current_name)
	if not source or not current:
		return True
	if SequenceMatcher(None, source, current).ratio() >= 0.55:
		return True
	stop_words = {
		"avec", "bosch", "dans", "equipement", "machine", "outil", "outils",
		"pour", "professional", "station", "table", "test",
	}
	source_tokens = {token for token in source.split() if len(token) >= 4 and token not in stop_words}
	current_tokens = {token for token in current.split() if len(token) >= 4 and token not in stop_words}
	if source_tokens.intersection(current_tokens):
		return True
	# A shared manufacturer is useful when the exact Item code also matches.
	manufacturers = {"bosch", "diatest", "mitutoyo", "sylvac", "tesa", "neriox", "skil"}
	return bool(set(source.split()).intersection(set(current.split())).intersection(manufacturers))


def parse_workbook_date(value):
	if isinstance(value, datetime):
		return value.date()
	if isinstance(value, date):
		return value
	if not isinstance(value, str):
		return None
	value = value.strip()
	for pattern in ("%d.%m.%Y", "%d.%m.%y", "%d/%m/%Y", "%Y-%m-%d"):
		try:
			return datetime.strptime(value, pattern).date()
		except ValueError:
			pass
	return None


def _item_code_candidates(value):
	value = _clean_value(value)
	if not value:
		return []
	return [part.strip() for part in value.split("/") if part.strip()]


def _normalize_name(value):
	value = unicodedata.normalize("NFKD", cstr(value or ""))
	value = "".join(character for character in value if not unicodedata.combining(character))
	return " ".join(re.findall(r"[a-z0-9]+", value.lower()))


def _normalize_maintenance_type(value):
	value = _normalize_name(value)
	if "calibr" in value:
		return "Calibration"
	if "repar" in value:
		return "Repair"
	if "modif" in value:
		return "Modification"
	if "control" in value or "inspect" in value:
		return "Inspection"
	if "nettoy" in value:
		return "Cleaning"
	if "mise en service" in value:
		return "Commissioning"
	if "maint" in value:
		return "Maintenance"
	return "Other"


def _employee_for_name(value):
	value = _clean_value(value)
	if not value:
		return None
	return frappe.db.get_value("Employee", {"employee_name": value, "status": "Active"}, "name")


def _clean_value(value):
	if value is None:
		return None
	if isinstance(value, float) and value.is_integer():
		value = int(value)
	value = cstr(value).strip()
	if value.lower() in ("n/a", "na", "non applicable", "/", "-"):
		return None
	return value or None


def _clean_header(value):
	return cstr(value or "").strip()


def _bucket_resolution_issue(report, sheet, row_number, row, reason):
	entry = {
		"sheet": sheet,
		"row": row_number,
		"item_code": _clean_value(row.get("Item code") or row.get("N°article")),
		"equipment": _clean_value(row.get("Name") or row.get("Equipement")),
		"reason": reason,
	}
	if reason == "code_name_conflict":
		report["conflicts"].append(entry)
	else:
		report["skipped"].append(entry)


def _resolve_workbook_path(file_path):
	file_path = cstr(file_path).strip()
	if not file_path:
		frappe.throw(_("A maintenance workbook path is required."))
	if file_path.startswith("/private/files/") or file_path.startswith("/files/"):
		file_path = frappe.get_site_path(file_path.lstrip("/"))
	path = os.path.abspath(file_path)
	if not os.path.isfile(path):
		frappe.throw(_("Maintenance workbook was not found: {0}").format(path))
	return path


def _maintenance_columns_exist():
	try:
		return all(frappe.db.has_column("Item", fieldname) for fieldname in (
			"tool_last_maintenance_date",
			"tool_next_maintenance_date",
			"tool_maintenance_status",
			"tool_open_maintenance_plans",
			"tool_overdue_maintenance_plans",
		))
	except Exception:
		return False
